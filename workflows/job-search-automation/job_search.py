#!/usr/bin/env python3
"""
Daily PM job search — appends new listings to Nick Application Tracker.xlsx.
Run via macOS LaunchAgent. Requires ANTHROPIC_API_KEY env var.
"""

import json
import logging
import os
import re
import subprocess
import time
from datetime import date
from pathlib import Path

import anthropic
import openpyxl
from openpyxl.styles import Font, PatternFill

# ── Paths ──────────────────────────────────────────────────────────────────────
BASE    = Path("/Users/nickgardone/Documents/Business/Career/Applications")
TRACKER = BASE / "Nick Application Tracker.xlsx"
LOG_FILE = BASE / "job_search.log"
SHEET   = "Job Opportunities"

# ── Colors (matched exactly from existing sheet) ───────────────────────────────
WHITE  = PatternFill(start_color="FFFFFFFF", end_color="FFFFFFFF", fill_type="solid")
GREEN  = PatternFill(start_color="FFE8F5E9", end_color="FFE8F5E9", fill_type="solid")  # $150k+ alt rows
BLUE   = PatternFill(start_color="FFD6E4F0", end_color="FFD6E4F0", fill_type="solid")  # not listed alt rows
ORANGE = PatternFill(start_color="FFFFF3E0", end_color="FFFFF3E0", fill_type="solid")  # below $150k

# ── Logging ────────────────────────────────────────────────────────────────────
logging.basicConfig(
    filename=LOG_FILE,
    level=logging.INFO,
    format="%(asctime)s  %(message)s",
    datefmt="%Y-%m-%d %H:%M",
)
log = logging.getLogger(__name__)

SEARCH_PROMPT = """
Search multiple job boards and return a JSON array of current Product Manager job listings.

CRITERIA:
- Roles to include: Senior Product Manager, Lead Product Manager, Staff Product Manager, Product Manager
- Roles to exclude: Principal Product Manager — do NOT include under any circumstances
- Location: "Remote – US" (fully remote, must be eligible for US-based employees) OR "Hybrid – Dallas, TX" (DFW metro, some in-office days)
- Do not include roles that are remote globally only with no US eligibility
- Salary target: $150,000+/year. Include roles with no listed salary. Flag roles with listed salary below $150k.
- Industry: Tech/SaaS preferred; any software-focused PM role is acceptable
- Company stage: any (startup, growth, public, enterprise)

Job boards to search: LinkedIn, Greenhouse (greenhouse.io), Lever (lever.co), Ashby (ashbyhq.com),
Remotive, Himalayas, We Work Remotely, Remote Rocketship, Built In, Indeed, Glassdoor, ZipRecruiter

Find at least 50 listings. Cast a wide net across multiple boards.

Return ONLY a valid JSON array — no preamble, no markdown fences, no explanation. Each element must have exactly these fields:
{
  "company": "Company Name",
  "title": "Exact job title as listed",
  "location_type": "Remote – US" or "Hybrid – Dallas, TX",
  "salary_range": "$X–$Y" or "—" if not listed,
  "salary_status": "✓ $150k+" or "⚠ Below $150k" or "—",
  "listing_date": "YYYY-MM-DD" or "—" if not found,
  "source": "Job board name",
  "apply_link": "https://...",
  "notes": "One sentence: industry, company type, or notable context"
}
"""


def notify(title: str, message: str) -> None:
    subprocess.run([
        "osascript", "-e",
        f'display notification "{message}" with title "{title}"'
    ], check=False)


def normalize(text: str) -> str:
    return re.sub(r"[^a-z0-9]", "", text.lower())


def load_existing(ws) -> tuple[set, set, int]:
    """Return (existing_urls, existing_company_title_keys, max_row_number)."""
    urls, keys, max_num = set(), set(), 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        num     = row[0]
        company = str(row[2] or "")
        title   = str(row[3] or "")
        url     = str(row[9] or "")
        if url.startswith("http"):
            urls.add(url.strip().rstrip("/"))
        if company and title:
            keys.add(normalize(company) + "|" + normalize(title))
        if isinstance(num, int) and num > max_num:
            max_num = num
    return urls, keys, max_num


def is_duplicate(listing: dict, urls: set, keys: set) -> bool:
    url = listing.get("apply_link", "").strip().rstrip("/")
    if url and url in urls:
        return True
    c = listing.get("company", "")
    t = listing.get("title", "")
    if c and t and normalize(c) + "|" + normalize(t) in keys:
        return True
    return False


def extract_listings(text: str) -> list[dict]:
    """Parse listings from Claude's response, recovering partial results if truncated."""
    # Try clean full-array parse first
    match = re.search(r"\[.*\]", text, re.DOTALL)
    if match:
        try:
            return json.loads(match.group())
        except json.JSONDecodeError:
            pass
    # Fall back: extract every complete {...} object individually
    listings = []
    for m in re.finditer(r"\{[^{}]*\}", text, re.DOTALL):
        try:
            obj = json.loads(m.group())
            if "company" in obj and "apply_link" in obj:
                listings.append(obj)
        except json.JSONDecodeError:
            continue
    if listings:
        log.warning(f"JSON was truncated — recovered {len(listings)} complete objects")
        return listings
    raise ValueError(f"Could not parse any listings. Raw excerpt: {text[:500]}")


def search_jobs(client: anthropic.Anthropic) -> list[dict]:
    for attempt in range(3):
        try:
            response = client.messages.create(
                model="claude-sonnet-4-6",
                max_tokens=16000,
                tools=[{"type": "web_search_20250305", "name": "web_search", "max_uses": 15}],
                messages=[{"role": "user", "content": SEARCH_PROMPT}],
            )
            text = "".join(b.text for b in response.content if hasattr(b, "text"))
            return extract_listings(text)
        except anthropic.RateLimitError as e:
            if attempt < 2:
                wait = 60 * (attempt + 1)
                log.warning(f"Rate limited — waiting {wait}s before retry {attempt + 2}/3")
                time.sleep(wait)
            else:
                raise


def find_listing_date(client: anthropic.Anthropic, company: str, title: str) -> str:
    """Search alternate sources to find a posting date for a listing that had none."""
    prompt = (
        f'Find the date this job was posted: "{title}" at {company}. '
        'Search LinkedIn, Indeed, Glassdoor, and the company careers page. '
        'Return ONLY a date in YYYY-MM-DD format, or "—" if you cannot find it. No other text.'
    )
    response = client.messages.create(
        model="claude-haiku-4-5-20251001",
        max_tokens=50,
        tools=[{"type": "web_search_20250305", "name": "web_search", "max_uses": 3}],
        messages=[{"role": "user", "content": prompt}],
    )
    text = "".join(b.text for b in response.content if hasattr(b, "text")).strip()
    m = re.search(r"\d{4}-\d{2}-\d{2}", text)
    return m.group() if m else "—"


def row_fill(salary_status: str, row_num: int) -> PatternFill:
    # Rows alternate: even data-index (row 2, 4, 6…) → white; odd → colored
    is_colored = (row_num - 2) % 2 == 1
    if "150k+" in str(salary_status):
        return GREEN if is_colored else WHITE
    elif "Below" in str(salary_status):
        return ORANGE
    else:
        return BLUE if is_colored else WHITE


def append_rows(ws, listings: list[dict], start_num: int, start_row: int) -> None:
    for i, item in enumerate(listings):
        r = start_row + i
        values = [
            start_num + i,
            date.today(),
            item.get("company", ""),
            item.get("title", ""),
            item.get("location_type", ""),
            item.get("salary_range", "—"),
            item.get("salary_status", "—"),
            item.get("listing_date", "—"),
            item.get("source", ""),
            item.get("apply_link", ""),
            item.get("notes", ""),
            "To Review",
        ]
        fill = row_fill(item.get("salary_status", ""), r)
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=r, column=col, value=val)
            cell.fill = fill
            if col == 2:
                cell.number_format = "MM/DD/YYYY"
            if col == 10 and val:
                cell.hyperlink = val
                cell.font = Font(color="0563C1", underline="single")


def main() -> None:
    log.info("── Job search started ──")

    api_key = os.environ.get("ANTHROPIC_API_KEY")
    if not api_key:
        log.error("ANTHROPIC_API_KEY not set — aborting.")
        return

    client = anthropic.Anthropic(api_key=api_key)

    wb = openpyxl.load_workbook(TRACKER)
    ws = wb[SHEET]
    existing_urls, existing_keys, max_num = load_existing(ws)
    log.info(
        f"Existing entries: {max_num}  "
        f"(dedup set: {len(existing_urls)} URLs, {len(existing_keys)} name+title pairs)"
    )

    try:
        raw = search_jobs(client)
        log.info(f"Claude returned {len(raw)} listings")
    except Exception as e:
        log.error(f"Search failed: {e}")
        notify("Job Search Failed ❌", f"Error: {str(e)[:80]}")
        return

    new_listings = []
    for item in raw:
        if is_duplicate(item, existing_urls, existing_keys):
            continue
        if item.get("listing_date", "—") == "—":
            try:
                item["listing_date"] = find_listing_date(
                    client, item.get("company", ""), item.get("title", "")
                )
            except Exception as e:
                log.warning(f"Date lookup failed for {item.get('company')}: {e}")
        new_listings.append(item)

    dupes = len(raw) - len(new_listings)
    log.info(f"New after dedup: {len(new_listings)}  Skipped as duplicates: {dupes}")

    if new_listings:
        next_row = ws.max_row + 1
        append_rows(ws, new_listings, max_num + 1, next_row)
        wb.save(TRACKER)
        log.info(f"Saved. Rows added: {next_row}–{next_row + len(new_listings) - 1}")

    log.info(f"── Done. {len(new_listings)} added, {dupes} duplicates skipped ──\n")
    notify(
        "Job Search Complete ✅",
        f"{len(new_listings)} new roles added, {dupes} duplicates skipped"
    )


if __name__ == "__main__":
    main()
